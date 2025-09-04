from typing import List, Dict, Any, Optional
import io
import csv
from math import sqrt
from reportlab.pdfgen import canvas
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Reuse common PDF helpers from export module
from .export import _mm_to_pt, _hex_to_color


def _draw_header(c: canvas.Canvas, title: str, subtitle: Optional[str] = None):
    c.setFillColor(_hex_to_color("#0F172A"))
    c.setFont("Helvetica-Bold", 14)
    c.drawString(_mm_to_pt(15), _mm_to_pt(287), title)
    if subtitle:
        c.setFont("Helvetica", 9)
        c.setFillColor(_hex_to_color("#334155"))
        c.drawString(_mm_to_pt(15), _mm_to_pt(282), subtitle)
    # top rule
    c.setStrokeColor(_hex_to_color("#CBD5E1"))
    c.setLineWidth(0.7)
    c.line(_mm_to_pt(12), _mm_to_pt(278), _mm_to_pt(198), _mm_to_pt(278))


def _draw_table_header(
    c: canvas.Canvas, x: float, y: float, columns: List[str], widths: List[float]
):
    c.setFillColor(_hex_to_color("#111827"))
    c.setFont("Helvetica-Bold", 9)
    cx = x
    for i, col in enumerate(columns):
        c.drawString(cx, y, col)
        cx += widths[i]
    # underline
    c.setStrokeColor(_hex_to_color("#94A3B8"))
    c.setLineWidth(0.5)
    c.line(x, y - 2, x + sum(widths), y - 2)


def _ensure_page_space(c: canvas.Canvas, cur_y: float, min_y: float, on_new_page):
    if cur_y < min_y:
        c.showPage()
        on_new_page()
        return True
    return False


def _fmt_dims_for_piece(p: Dict[str, Any]) -> str:
    # For rectangles: WxH mm; for polygons: bbox WxH and edge list
    if p.get("polygon"):
        pts = p["polygon"]
        xs = [pt[0] for pt in pts]
        ys = [pt[1] for pt in pts]
        bw = int(round(max(xs) - min(xs)))
        bh = int(round(max(ys) - min(ys)))
        edges = []
        for i in range(len(pts)):
            x0, y0 = pts[i]
            x1, y1 = pts[(i + 1) % len(pts)]
            L = int(round(sqrt((x1 - x0) ** 2 + (y1 - y0) ** 2)))
            edges.append(str(L))
        return f"bbox {bw}×{bh} mm; edges {', '.join(edges)} mm"
    else:
        w = int(round(p.get("width", 0)))
        h = int(round(p.get("height", 0)))
        # include four edge dims like visuals (two of each)
        return f"{w}×{h} mm (edges: {w}, {h}, {w}, {h})"


def _wrap_text(
    c: canvas.Canvas,
    text: str,
    max_width_pt: float,
    font: str = "Helvetica",
    size: float = 9.0,
) -> List[str]:
    """Wrap text to fit within max_width_pt using canvas.stringWidth.

    Splits on spaces and commas while preserving separators, and falls back to
    character-level breaking if a single token exceeds the width.
    """
    if not text:
        return [""]
    tokens = re.split(r"(,\s*|\s+)", text)
    lines: List[str] = []
    cur = ""
    for tok in tokens:
        if tok is None:
            continue
        tentative = cur + tok
        w = c.stringWidth(tentative, font, size)
        if w <= max_width_pt or not cur:
            cur = tentative
            continue
        # commit current line and start new
        lines.append(cur.rstrip())
        # if token alone is too wide, break it by characters
        if c.stringWidth(tok, font, size) > max_width_pt:
            piece = ""
            for ch in tok:
                if c.stringWidth(piece + ch, font, size) <= max_width_pt or not piece:
                    piece += ch
                else:
                    lines.append(piece)
                    piece = ch
            cur = piece
        else:
            cur = tok
    if cur:
        lines.append(cur.rstrip())
    return lines


def cutsheet_to_pdf_bytes(
    job_name: str,
    cabinets: List[Dict[str, Any]],
    pieces_by_cab: Dict[str, List[Dict[str, Any]]],
    colours_by_id: Optional[Dict[str, Dict[str, Any]]] = None,
    title: str = "Cut Sheet",
) -> bytes:
    """Render a grouped cut sheet to a multi-page PDF (A4 portrait).

    - Groups by cabinet; each cabinet gets a section header and a table of pieces.
    - Columns: Name, Type, Material, Dimensions.
    - No external system deps beyond reportlab.
    """
    buf = io.BytesIO()
    # A4 portrait: 210 x 297 mm
    c = canvas.Canvas(buf, pagesize=(_mm_to_pt(210), _mm_to_pt(297)))
    safe_right = _mm_to_pt(198)
    safe_top = _mm_to_pt(278)
    safe_bottom = _mm_to_pt(15)

    def start_page():
        _draw_header(c, f"{title} — {job_name}")

    start_page()

    # Table layout
    x0 = _mm_to_pt(15)
    col_widths = [
        _mm_to_pt(70),  # Name
        _mm_to_pt(25),  # Type
        _mm_to_pt(80),  # Dimensions
    ]
    columns = ["Piece", "Type", "Dimensions"]

    cur_y = safe_top

    def new_page_if_needed():
        nonlocal cur_y
        if _ensure_page_space(c, cur_y, safe_bottom + _mm_to_pt(30), start_page):
            cur_y = safe_top

    for cab in cabinets:
        cab_name = cab.get("name") or cab.get("id") or "Cabinet"
        # Cabinet header
        new_page_if_needed()
        c.setFillColor(_hex_to_color("#111827"))
        c.setFont("Helvetica-Bold", 12)
        cur_y -= _mm_to_pt(8)
        c.drawString(x0, cur_y, f"Cabinet: {cab_name}")
        cur_y -= _mm_to_pt(3)

        # Table header
        _draw_table_header(c, x0, cur_y, columns, col_widths)
        cur_y -= _mm_to_pt(6)

        # Rows
        rows = pieces_by_cab.get(cab.get("id"), [])
        for p in rows:
            new_page_if_needed()
            c.setFont("Helvetica", 9)
            c.setFillColor(_hex_to_color("#0F172A"))
            name_x = x0
            type_x = x0 + col_widths[0]
            dims_x = type_x + col_widths[1]

            # Left cells (single line)
            name = str(p.get("name") or p.get("id") or "")
            c.drawString(name_x, cur_y, name[:48])
            typ = "Poly" if p.get("polygon") else "Rect"
            c.drawString(type_x, cur_y, typ)

            # Dimensions with wrapping within the remaining column width
            dims_text = _fmt_dims_for_piece(p)
            dims_width = col_widths[2]
            lines = _wrap_text(c, dims_text, dims_width, font="Helvetica", size=9)
            line_step = _mm_to_pt(5)
            # Draw first line at current y, subsequent lines below
            for i, line in enumerate(lines):
                if i > 0:
                    cur_y -= line_step
                    new_page_if_needed()
                    c.setFont("Helvetica", 9)
                    c.setFillColor(_hex_to_color("#0F172A"))
                c.drawString(dims_x, cur_y, line)
            # Space after row
            cur_y -= _mm_to_pt(5)

        # Spacer after cabinet
        cur_y -= _mm_to_pt(3)

    # Footer
    c.setFont("Helvetica", 8)
    c.setFillColor(_hex_to_color("#64748B"))
    c.drawRightString(safe_right, _mm_to_pt(8), "Generated by Stroptimise")

    c.save()
    return buf.getvalue()


def cutsheet_to_csv_bytes(
    cabinets: List[Dict[str, Any]],
    pieces_by_cab: Dict[str, List[Dict[str, Any]]],
) -> bytes:
    """Produce a CSV with columns: Cabinet, Piece, Type, Dimensions."""
    sio = io.StringIO(newline="")
    writer = csv.writer(sio)
    writer.writerow(["Cabinet", "Piece", "Type", "Dimensions"])
    # preserve cabinet order as provided
    for cab in cabinets:
        cab_name = cab.get("name") or cab.get("id") or "Cabinet"
        for p in pieces_by_cab.get(cab.get("id"), []):
            name = str(p.get("name") or p.get("id") or "")
            typ = "Poly" if p.get("polygon") else "Rect"
            dims = _fmt_dims_for_piece(p)
            writer.writerow([cab_name, name, typ, dims])
    return sio.getvalue().encode("utf-8-sig")


def cutsheet_to_xlsx_bytes(
    cabinets: List[Dict[str, Any]],
    pieces_by_cab: Dict[str, List[Dict[str, Any]]],
) -> bytes:
    """Produce an XLSX with columns: Cabinet, Piece, Type, Dimensions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cut Sheet"
    headers = ["Cabinet", "Piece", "Type", "Dimensions"]
    ws.append(headers)
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    # Rows
    for cab in cabinets:
        cab_name = cab.get("name") or cab.get("id") or "Cabinet"
        for p in pieces_by_cab.get(cab.get("id"), []):
            name = str(p.get("name") or p.get("id") or "")
            typ = "Poly" if p.get("polygon") else "Rect"
            dims = _fmt_dims_for_piece(p)
            ws.append([cab_name, name, typ, dims])
    # Freeze header and set widths
    ws.freeze_panes = "A2"
    widths = [18, 22, 10, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Output to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ===== Sheet-grouped variants (use latest placement results) =====


def cutsheet_by_sheet_to_pdf_bytes(
    job_name: str,
    sheets: List[Dict[str, Any]],  # [{id, name, width, height, index}]
    rows_by_sheet: Dict[
        str, List[Dict[str, Any]]
    ],  # sheet_id -> [{cabinet_name, name, polygon|width/height}]
    title: str = "Cut Sheet",
) -> bytes:
    buf = io.BytesIO()
    # A4 portrait
    c = canvas.Canvas(buf, pagesize=(_mm_to_pt(210), _mm_to_pt(297)))
    safe_right = _mm_to_pt(198)
    safe_top = _mm_to_pt(278)
    safe_bottom = _mm_to_pt(15)

    def start_page():
        _draw_header(c, f"{title} — {job_name}")

    start_page()

    # Columns: Cabinet, Piece, Type, Dimensions
    x0 = _mm_to_pt(15)
    col_widths = [
        _mm_to_pt(35),  # Cabinet
        _mm_to_pt(55),  # Piece
        _mm_to_pt(20),  # Type
        _mm_to_pt(60),  # Dimensions
    ]
    columns = ["Cabinet", "Piece", "Type", "Dimensions"]

    cur_y = safe_top

    def new_page_if_needed():
        nonlocal cur_y
        if _ensure_page_space(c, cur_y, safe_bottom + _mm_to_pt(30), start_page):
            cur_y = safe_top

    for sh in sorted(sheets, key=lambda s: s.get("index", 0)):
        # Section header for sheet
        new_page_if_needed()
        c.setFillColor(_hex_to_color("#111827"))
        c.setFont("Helvetica-Bold", 12)
        cur_y -= _mm_to_pt(8)
        sh_idx = sh.get("index", 0)
        sh_name = sh.get("name")
        if sh_name:
            sh_title = f"Sheet {sh_idx} — {sh_name}"
        else:
            sh_title = f"Sheet {sh_idx}"
        # include size hint
        sz = ""
        if sh.get("width") and sh.get("height"):
            sz = f"  ({sh['width']}×{sh['height']} mm)"
        c.drawString(x0, cur_y, f"{sh_title}{sz}")
        cur_y -= _mm_to_pt(3)

        # Table header
        _draw_table_header(c, x0, cur_y, columns, col_widths)
        cur_y -= _mm_to_pt(6)

        rows = rows_by_sheet.get(sh.get("id"), [])
        for p in rows:
            new_page_if_needed()
            c.setFont("Helvetica", 9)
            c.setFillColor(_hex_to_color("#0F172A"))
            cab_x = x0
            piece_x = cab_x + col_widths[0]
            type_x = piece_x + col_widths[1]
            dims_x = type_x + col_widths[2]

            c.drawString(cab_x, cur_y, str(p.get("cabinet_name") or "")[:24])
            c.drawString(piece_x, cur_y, str(p.get("name") or p.get("id") or "")[:40])
            typ = "Poly" if p.get("polygon") else "Rect"
            c.drawString(type_x, cur_y, typ)

            dims_text = _fmt_dims_for_piece(p)
            dims_width = col_widths[3]
            lines = _wrap_text(c, dims_text, dims_width, font="Helvetica", size=9)
            line_step = _mm_to_pt(5)
            for i, line in enumerate(lines):
                if i > 0:
                    cur_y -= line_step
                    new_page_if_needed()
                    c.setFont("Helvetica", 9)
                    c.setFillColor(_hex_to_color("#0F172A"))
                c.drawString(dims_x, cur_y, line)
            cur_y -= _mm_to_pt(5)

        cur_y -= _mm_to_pt(3)

    c.setFont("Helvetica", 8)
    c.setFillColor(_hex_to_color("#64748B"))
    c.drawRightString(safe_right, _mm_to_pt(8), "Generated by Stroptimise")
    c.save()
    return buf.getvalue()


def cutsheet_by_sheet_to_csv_bytes(
    sheets: List[Dict[str, Any]],
    rows_by_sheet: Dict[str, List[Dict[str, Any]]],
) -> bytes:
    sio = io.StringIO(newline="")
    writer = csv.writer(sio)
    writer.writerow(["Sheet #", "Sheet", "Cabinet", "Piece", "Type", "Dimensions"])
    for sh in sorted(sheets, key=lambda s: s.get("index", 0)):
        sh_idx = sh.get("index", 0)
        sh_label = sh.get("name") or f"Sheet {sh_idx}"
        for p in rows_by_sheet.get(sh.get("id"), []):
            name = str(p.get("name") or p.get("id") or "")
            typ = "Poly" if p.get("polygon") else "Rect"
            dims = _fmt_dims_for_piece(p)
            writer.writerow(
                [sh_idx, sh_label, p.get("cabinet_name") or "", name, typ, dims]
            )
    return sio.getvalue().encode("utf-8-sig")


def cutsheet_by_sheet_to_xlsx_bytes(
    sheets: List[Dict[str, Any]],
    rows_by_sheet: Dict[str, List[Dict[str, Any]]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cut Sheet"
    headers = ["Sheet #", "Sheet", "Cabinet", "Piece", "Type", "Dimensions"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for sh in sorted(sheets, key=lambda s: s.get("index", 0)):
        sh_idx = sh.get("index", 0)
        sh_label = sh.get("name") or f"Sheet {sh_idx}"
        for p in rows_by_sheet.get(sh.get("id"), []):
            name = str(p.get("name") or p.get("id") or "")
            typ = "Poly" if p.get("polygon") else "Rect"
            dims = _fmt_dims_for_piece(p)
            ws.append([sh_idx, sh_label, p.get("cabinet_name") or "", name, typ, dims])
    ws.freeze_panes = "A2"
    widths = [10, 18, 18, 24, 10, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
